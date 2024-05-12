import os
import time
import pythoncom
from pprint import pprint as pp
import re
import win32com
import win32com.client as win32 
from win32com.client import constants
from pptx import Presentation
import pandas as pd
import tkinter as tk
from tkinter import *
from tkinter import messagebox, filedialog
from tkinter.filedialog import askopenfile
from tkinter.scrolledtext import ScrolledText
from tkinter import ttk
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl.utils.exceptions as xlutils
from openpyxl.utils import rows_from_range, range_boundaries
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.drawing.image import Image
import glob
import xlwt
import xlrd
from xlsxwriter import utility
import numpy as np
import io
#%matplotlib inline
from xlwings.constants import DeleteShiftDirection
import logging
#import traceback
import sys
from PyPDF2 import PdfFileReader, PdfFileWriter,PdfFileMerger
import datetime
import traceback
import docx
from docx.shared import Pt
import xlwings as xw
import calendar
import pyxlsb
from pyxlsb import open_workbook as open_xlsb
import warnings
warnings.filterwarnings('ignore')
from threading import Thread

class Console():
    def __init__(self):
        self.Main = Tk()
        self.Window_Config()
        self.Memory()
        self.load_Contents()

    #####################################################
    ################## Actual Console ###################
    #####################################################

    def Excuse_Buffer(self, duration):
        time.sleep(duration)
    
    def PTrace(self, T):
        return "\t"+"\n\t".join([line for line in T.split('\n')])

    def Window_Config(self):
        self.Main.wm_title('ARC')
        width = 850
        height = 500
        screen_width = self.Main.winfo_screenwidth()
        screen_height = self.Main.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        self.Main.geometry('%dx%d+%d+%d' % (width, height, x, y))
        self.Main.grid_columnconfigure(1,minsize=70,weight=1)
        self.Main.configure(bg='white')
        self.Main.resizable(False, False)
    
    def load_Contents(self):
        self.load_CFX_HEADER()
        self.load_BUTTON_FRAME()
        self.load_Console_Frame()

    def load_CFX_HEADER(self):
        self.Title = Label(self.Main, text='Reporting Automation Console', height=1, font=('Ariel', 20, 'bold'), fg='white', bg='red4')
        self.Title.grid(row=0, column=0, padx=10, pady=5, columnspan=10, sticky='ew')

    def load_BUTTON_FRAME(self):
        self.Button_Frame = Frame(self.Main, width=650, height=100, background='lightgray', highlightbackground="black", highlightthickness=1)
        self.Button_Frame.grid(row=1, column=0, padx=10, pady=5, columnspan=10, sticky='ew')
        self.Button_Frame.pack_propagate(False)
        self.Process1 = Button(self.Button_Frame, text='Excel to PPT', width=12, command=lambda: self.Show_Status('Excel to PPT'), height=1, font=('Ariel', 12, 'bold'), fg='white', bg='red4')
        self.Process1.place(x=21, y=30)
        self.Process2 = Button(self.Button_Frame, text='Control', width=12, command=lambda: self.Show_Status('Control/Validation'), height=1, font=('Ariel', 12, 'bold'), fg='white', bg='red4')
        self.Process2.place(x=184, y=30)
        self.Process3 = Button(self.Button_Frame, text='More1', width=12, command=lambda: self.Show_Status(''), height=1, font=('Ariel', 12, 'bold'), fg='white', bg='red4')
        self.Process3.place(x=348, y=30)
        self.Process4 = Button(self.Button_Frame, text='More2', width=12, command=lambda: self.Show_Status(''), height=1, font=('Ariel', 12, 'bold'), fg='white', bg='red4')
        self.Process4.place(x=513, y=30)
        self.Process5 = Button(self.Button_Frame, text='More3', width=12, command=lambda: self.Show_Status(''), height=1, font=('Ariel', 12, 'bold'), fg='white', bg='red4')
        self.Process5.place(x=675, y=30)
    
    def load_Console_Frame(self):
        self.Console_Frame = Frame(self.Main, width=650, height=327, background='lightgray', highlightbackground="black", highlightthickness=1)
        self.Console_Frame.grid(row=2, column=0, padx=10, pady=5, columnspan=10, sticky='ew')

        self.Console_Out = ScrolledText(self.Console_Frame, width=200,  height= 20, )
        self.Console_Out.pack(fill=BOTH, side=LEFT, expand=True)

        self.Curr_Data = '''\tHello there.
        This is a standard console.
        Your process states, outputs and error tracebacks will appear here.
        Do to panic. This Window will keep you updated.'''
        
        self.Update_Console(self.Curr_Data)

    def Update_Console(self, WData):
        self.Curr_Data += WData
        WData += '\n\t----------------------------------------------------------------------\n'
        self.Console_Out.insert('end', WData)

    def Show_Status(self, Process):
        self.Update_Console('\n\t*Status enquiry called by user for '+Process+' process.*')
        if Process in self.Processed:
            messagebox.showinfo('Status', Process+' is completed successfully. Check Results.')
            return
        if Process in self.ToBeProcessed:
            messagebox.showinfo('Status', Process+' is yet not started. Please wait.')
            return
        if self.Curr_Status == Process:
            messagebox.showwarning('Status', Process+' is currently being executed. Please deal with Excel/System dialogues outside the tool.')
            return
        messagebox.showerror('Status', Process+' not defined in Memory.')
        
    
    def Memory(self):
        self.Curr_Status = None
        self.Processed = []
        self.ToBeProcessed = ['Excel to PPT', 'Control/Validation']

    def Start_Engine(self):
        self.Main.mainloop()
    
    #####################################################
    ################# Actual Functions ##################
    #####################################################

    def Start_Process(self):
        try:
            self.Excuse_Buffer(5)
            self.Exceltoppt()
            self.Control_Validation_Revamped()
        except Exception as e:
            self.Update_Console('\tError Test : '+str(e))
    
    def titleToNumber(self, s): 
        result = 0; 
        for B in range(len(s)): 
            result *= 26; 
            result += ord(s[B]) - ord('A') + 1; 
        return result

    def ret_Range(self, Col_Num):
        Col_List = []
        Ranges = Col_Num.split(',')
        for R in Ranges:
            R = R.replace(' ', '')
            if ':' in R:
                StartR, EndR = R.split(':')
                StartR, EndR = self.titleToNumber(StartR), self.titleToNumber(EndR)
                for i in range(StartR, EndR+1):
                    Col_List.append(utility.xl_col_to_name(i-1))
            else:
                Col_List.append(R)
        return Col_List

    def Exceltoppt(self): 
        try:
            self.xl = win32.Dispatch('Excel.Application', pythoncom.CoInitialize())
            self.xl.Workbooks.Add()
            self.xl.DisplayAlerts= False
            self.xl.Visible= False   
            self.xl.Application.Calculation = -4135
            xl_path = r'C:\Users\45366682\Desktop\AutoTemplate\AutoValidate\HBCE Input Template.xlsx'
            wb_obj = openpyxl.load_workbook(xl_path)
            sheet_obj = wb_obj["Ex2PPT"]        
            c= sheet_obj.max_column 
            r = sheet_obj.max_row  
            PPT_path =  (sheet_obj.cell(row = 3, column = 1)).value
            PPT_outputpath =  (sheet_obj.cell(row = 3, column = 2)).value
            PPTApp = win32.Dispatch('PowerPoint.Application')
            PPTApp.Visible = True
            ppt1 = PPTApp.Presentations.Open(PPT_path) 
            start_time = time.time() 
            for i in range(7,r+1):
                try:
                    excelpath = (sheet_obj.cell(row = i, column = 1)).value
                    worksheet = (sheet_obj.cell(row = i, column = 2)).value
                    excelrange = (sheet_obj.cell(row = i, column = 3)).value
                    slideno = int((sheet_obj.cell(row = i, column = 4)).value)
                    slidehg = int((sheet_obj.cell(row = i, column = 5)).value)
                    slidewd = int((sheet_obj.cell(row = i, column = 6)).value)
                    slideleft = int((sheet_obj.cell(row = i, column = 7)).value)
                    slidetop = int((sheet_obj.cell(row = i, column = 8)).value)
                    password=(sheet_obj.cell(row = i, column = 9)).value
                    #linked_version = int((sheet_obj.cell(row = i, column = 9)).value) 
                
                    if (excelpath == None):
                        pass
                    else:
                        wb2 = self.xl.Workbooks.Open(excelpath, 0, True, Password='password', WriteResPassword='')
                        wb2.Password = ''
                        PPTSlide = ppt1.Slides(slideno)
                        xl_worksheet = wb2.Worksheets(worksheet)
                        xl_range = xl_worksheet.Range(excelrange)
                        xl_range.Copy()
                        PPTSlide.Shapes.PasteSpecial(DataType = 1, Link = True) 
                        PPTShape = PPTSlide.Shapes(PPTSlide.Shapes.Count)
                        PPTShape.Left = slideleft
                        PPTShape.Top = slidetop
                        PPTShape.Height = slidehg
                        PPTShape.Width = slidewd 
                    pass               
                except Exception as e:
                    self.Update_Console('\t'+'Excel to PPT : '+'\n\t'+ str(e)+self.PTrace(traceback.format_exc()))

            start_time1 = time.time() 
            ppt1.SaveAs(PPT_outputpath) 

            dif_time = "{:.2f}".format( start_time1 - start_time)
            self.Update_Console('\n\tExcel to PPT completed successfully.') 
        except Exception as e: 
            self.Update_Console('\t'+'Excel to PPT : '+'\n\t'+str(e)+'\n'+self.PTrace(traceback.format_exc()))

    def Control_Validation_Revamped(self):
        try:
            Inp_File = r'C:\Users\45366682\Desktop\AutoTemplate\AutoValidate\HBCE Input Template.xlsx'
            Gen_DF = pd.read_excel(Inp_File, sheet_name='Control-Range', engine='openpyxl')
            Out_File = Gen_DF.iloc[0]['Output']
            Central_Mapper = {}
            #print(Gen_DF)
            for i in range(Gen_DF.shape[0]):
                Curr_Range = Gen_DF.iloc[i]
                Curr_File_Loc = Curr_Range['File Name']
                Curr_Sheet_Name = Curr_Range['Sheet Name']
                Curr_Range = Curr_Range['Range']
                if Curr_File_Loc not in Central_Mapper:
                    Central_Mapper[Curr_File_Loc] = {}
                Excel_Mapper = Central_Mapper[Curr_File_Loc][Curr_Sheet_Name] = {}
                StartR, EndR = Curr_Range.split(':')
                StartR_Col, EndR_Col = '', ''
                for C in StartR:
                    if C.isalpha():
                        StartR_Col+=C
                for C in EndR:
                    if C.isalpha():
                        EndR_Col+=C
                Header = [Col for Col in self.ret_Range(StartR_Col+':'+EndR_Col)]
                Curr_File = xw.Book(Curr_File_Loc)
                Curr_Sheet = Curr_File.sheets(Curr_Sheet_Name)
                Curr_DF_Val = pd.DataFrame(Curr_Sheet.range(Curr_Range).value)
                Curr_DF_Formula = pd.DataFrame(Curr_Sheet.range(Curr_Range).formula)
                #print(Curr_DF_Formula,Curr_DF_Val)
                try:
    
                    Curr_DF_Formula = Curr_DF_Formula.set_axis(Header, axis=1)
                except:
                    #Curr_DF_Val = Curr_DF_Val.set_axis(Header, axis=0)
                    Curr_DF_Formula = Curr_DF_Formula.set_axis(Header, axis=0)
                    Curr_DF_Formula = Curr_DF_Formula.T
    
                try:
                    Curr_DF_Val = Curr_DF_Val.set_axis(Header, axis=1)
                except:
                    Curr_DF_Val = Curr_DF_Val.set_axis(Header, axis=0)
                    Curr_DF_Val = Curr_DF_Val.T
                for j in range(Curr_DF_Val.shape[0]):
                    Val_Line = Curr_DF_Val.iloc[j]
                    Formula_Line = Curr_DF_Formula.iloc[j]
                    if Val_Line[Header[0]] == None:
                        continue
                    Control_Label = Val_Line[Header[0]]+f' [{StartR_Col+str(int(StartR.strip(StartR_Col))+j)}]'
                    Line_Mapper = Excel_Mapper[Control_Label] = {
                        'isVar' : False,
                        'VarCols' : [],
                        'isErr' : False,
                        'ErrCols' : []
                    }
                    for H in Header[1:]:
                        Value = Val_Line[H]
                        Formula = Formula_Line[H]
                        if Formula == '':
                            continue
                        try:
                            if round(Value, 2) != 0:
                                Line_Mapper['isVar'] = True
                                Line_Mapper['VarCols'].append(H)
                        except:
                            Line_Mapper['isErr'] = True
                            Line_Mapper['ErrCols'].append(H)
    
            FReport = xw.Book()
            Sht_cnt = 1
            Clearance = []
            # print(Central_Mapper)
            for File in Central_Mapper:
                Fname = File.split('\\')[-1]
                FReport.sheets('Sheet'+str(Sht_cnt)).api.Name = str(Sht_cnt)
                FSheets = FReport.sheets(str(Sht_cnt))
                Sht_cnt += 1
                FReport.sheets.add()
                Sheet_Mapper = Central_Mapper[File]
                FSheets.range('A1').value = ['File Name', Fname]
                FSheets.range('B1:C1').merge()
                FSheets.range('A1:C1').color = (255, 0, 0)
                FSheets.range("A1:C1").font.color = "#ffffff"
                FSheets.range("A1:C1").font.bold = True
                FSheets.range("A1:C1").font.size = 13
                FSheets.range('A1:C1').api.Borders.Weight = 3
                Start_Row = 2
    
                for Sheet in Sheet_Mapper:
                    Body = [['Sheet Name', Sheet, ''], ['Control Label', 'Variance Detected', 'Errors Detected']]
                    Excel_Mapper = Sheet_Mapper[Sheet]
                    for Control in Excel_Mapper:
                        Line_Mapper = Excel_Mapper[Control]
                        if not Line_Mapper['isErr'] and not Line_Mapper['isVar']:
                            continue
                        Row_Content = ['']*3
                        Row_Content[0] = Control
                        if Line_Mapper['isErr']:
                            Row_Content[2] = ", ".join(Line_Mapper['ErrCols'])
                        if Line_Mapper['isVar']:
                            Row_Content[1] = ", ".join(Line_Mapper['VarCols'])
                        Body.append(Row_Content)
                    Sheet_Table_Size = len(Body)
                    if Sheet_Table_Size == 2:
                        Clearance.append([File, Sheet])
                        continue
                    FSheets.range('A'+str(Start_Row)).value = Body
                    FSheets.range('B'+str(Start_Row)+':C'+str(Start_Row)).merge()
                    FSheets.range('A'+str(Start_Row)+':C'+str(Start_Row+1)).font.bold = True
                    FSheets.range('A'+str(Start_Row)+':C'+str(Start_Row+1)).font.size = 12
                    FSheets.range('A'+str(Start_Row)+':C'+str(Start_Row+1)).color = (0, 125, 125)
                    FSheets.range('A'+str(Start_Row)+':C'+str(Start_Row+1)).font.color = "#ffffff"
                    FSheets.range('A'+str(Start_Row+2)+':A'+str(Start_Row+Sheet_Table_Size-1)).color = (255, 255, 0)
                    FSheets.range('A'+str(Start_Row)+':C'+str(Start_Row+Sheet_Table_Size-1)).api.Borders.Weight = 2
                    Start_Row += Sheet_Table_Size + 4
    
                FSheets.range('A1:C'+str(Start_Row)).autofit()
                FSheets.range('A1:C'+str(Start_Row)).api.HorizontalAlignment = xw.constants.HAlign.xlHAlignCenter
                FSheets.range('A1:C'+str(Start_Row)).api.VerticalAlignment = xw.constants.VAlign.xlVAlignCenter
                time.sleep(2)
            FSheets = FReport.sheets('Sheet'+str(Sht_cnt))
            FSheets.delete()
            # FReport.save(Out_File)
            self.Update_Console('\t'+'Excel to PPT : '+'Completed successfully.')
        except Exception as e:
            self.Update_Console('\t'+'Excel to PPT : '+'\n\t'+str(e)+'\n'+self.PTrace(traceback.format_exc()))


if __name__ == "__main__":
    Instance = Console()
    Thread(target=Instance.Start_Process).start()
    Instance.Start_Engine()